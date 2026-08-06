"""Aplicação de bordas em células da planilha Excel."""

from openpyxl.worksheet.worksheet import Worksheet

from src.formatacao.estilos import BORDA_THIN, COLUNAS_BORDA_SEMPRE


def aplicar_bordas(ws: Worksheet, indices_col: dict) -> None:
    """
    Aplica bordas finas nas células com conteúdo.
    
    Algumas colunas (COLUNAS_BORDA_SEMPRE) recebem borda mesmo vazias.
    
    Args:
        ws: Worksheet do openpyxl
        indices_col: Dicionário mapeando nome da coluna -> índice
    """
    borda_sempre = {indices_col[c] for c in COLUNAS_BORDA_SEMPRE if c in indices_col}
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            tem_conteudo = cell.value is not None and str(cell.value).strip() != ""
            forcada = cell.column in borda_sempre and cell.row > 1
            if tem_conteudo or forcada:
                cell.border = BORDA_THIN
