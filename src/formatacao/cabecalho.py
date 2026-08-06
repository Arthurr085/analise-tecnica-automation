"""Formatação do cabeçalho da planilha Excel."""

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet

from src.formatacao.estilos import COR_CABECALHO


def formatar_cabecalho(ws: Worksheet) -> None:
    """
    Aplica formatação visual no cabeçalho (primeira linha).
    
    - Fundo azul
    - Texto branco e negrito
    - Alinhamento centralizado
    """
    fill = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
