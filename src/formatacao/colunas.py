"""Formatação de colunas da planilha Excel."""

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.formatacao.estilos import COLUNAS_CENTRO


def alinhar_centro(ws: Worksheet, indices_col: dict) -> None:
    """
    Centraliza o conteúdo das colunas especificadas em COLUNAS_CENTRO.
    
    Args:
        ws: Worksheet do openpyxl
        indices_col: Dicionário mapeando nome da coluna -> índice
    """
    colunas = {indices_col[c] for c in COLUNAS_CENTRO if c in indices_col}
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in colunas:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def ajustar_largura_colunas(ws: Worksheet) -> None:
    """
    Ajusta automaticamente a largura das colunas baseado no conteúdo.
    
    Largura mínima: 10 caracteres
    Largura máxima: 60 caracteres
    """
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)
