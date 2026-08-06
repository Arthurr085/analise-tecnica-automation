"""Módulo formatacao - Formatação visual de planilhas Excel."""

from openpyxl import load_workbook

from src.formatacao.cabecalho import formatar_cabecalho
from src.formatacao.bordas import aplicar_bordas
from src.formatacao.colunas import alinhar_centro, ajustar_largura_colunas


def formatar_planilha(caminho: str) -> None:
    """Aplica toda a formatação visual na planilha gerada."""
    wb = load_workbook(caminho)
    ws = wb.active

    indices_col = {cell.value: cell.column for cell in ws[1]}

    formatar_cabecalho(ws)
    aplicar_bordas(ws, indices_col)
    alinhar_centro(ws, indices_col)
    ajustar_largura_colunas(ws)
    ws.auto_filter.ref = ws.dimensions

    wb.save(caminho)


__all__ = ["formatar_planilha"]
