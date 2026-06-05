from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

_COR_CABECALHO = "8EAADB"

_BORDA_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_COLUNAS_CENTRO = {
    "Demanda",
    "Analista/Programador",
    "Objeto (Envolvimento na demanda)",
    "Programa Alterado",
    "Versão Pacote Atualização",
}

_COLUNAS_BORDA_SEMPRE = {"Demandas Pendentes"}


def formatar_planilha(caminho: str) -> None:
    """Aplica toda a formatação visual na planilha gerada."""
    wb = load_workbook(caminho)
    ws = wb.active

    indices_col = {cell.value: cell.column for cell in ws[1]}

    _formatar_cabecalho(ws)
    _aplicar_bordas(ws, indices_col)
    _alinhar_centro(ws, indices_col)
    _ajustar_largura_colunas(ws)
    ws.auto_filter.ref = ws.dimensions

    wb.save(caminho)


def _formatar_cabecalho(ws) -> None:
    fill = PatternFill(start_color=_COR_CABECALHO, end_color=_COR_CABECALHO, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _aplicar_bordas(ws, indices_col: dict) -> None:
    borda_sempre = {indices_col[c] for c in _COLUNAS_BORDA_SEMPRE if c in indices_col}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            tem_conteudo = cell.value is not None and str(cell.value).strip() != ""
            forcada = cell.column in borda_sempre and cell.row > 1
            if tem_conteudo or forcada:
                cell.border = _BORDA_THIN


def _alinhar_centro(ws, indices_col: dict) -> None:
    colunas = {indices_col[c] for c in _COLUNAS_CENTRO if c in indices_col}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in colunas:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def _ajustar_largura_colunas(ws) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)
